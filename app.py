from flask import Flask, render_template, request, jsonify
from db import get_connection

app = Flask(__name__)

@app.route("/")
def index():
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("SELECT country_id, country_name FROM dim_countries ORDER BY country_name")
    countries = cur.fetchall()
    cur.execute("SELECT indicator_id, indicator_name FROM dim_indicators ORDER BY indicator_id")
    indicators = cur.fetchall()
    cur.close()
    conn.close()
    return render_template("index.html", countries=countries, indicators=indicators)

@app.route("/data")
def data():
    country = request.args.get("country", "VNM")
    indicator_id = request.args.get("indicator", 1)
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("""
        SELECT t.year_id, f.value, t.is_crisis_year
        FROM fact_economic_data f
        JOIN dim_time t ON f.year_id = t.year_id
        WHERE f.country_id = %s AND f.indicator_id = %s
        ORDER BY t.year_id
    """, (country, indicator_id))
    rows = cur.fetchall()

    cur.execute("""
        SELECT definition FROM dim_indicators WHERE indicator_id = %s
    """, (indicator_id,))
    definition = cur.fetchone()

    cur.close()
    conn.close()
    return jsonify({
        "chartData": [{"year": r[0], "value": float(r[1]), "crisis": r[2]} for r in rows],
        "definition": definition[0] if definition else ""
    })

@app.route("/country/<country_id>")
def country_detail(country_id):
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("""
        SELECT c.country_name, r.region_name
        FROM dim_countries c
        JOIN dim_regions r ON c.region_id = r.region_id
        WHERE c.country_id = %s
    """, (country_id,))
    country = cur.fetchone()
    cur.execute("""
        SELECT i.indicator_name, f.year_id, f.value, u.unit_name
        FROM fact_economic_data f
        JOIN dim_indicators i ON f.indicator_id = i.indicator_id
        JOIN dim_units u ON i.unit_id = u.unit_id
        WHERE f.country_id = %s
        ORDER BY i.indicator_id, f.year_id
    """, (country_id,))
    rows = cur.fetchall()
    cur.close()
    conn.close()
    return render_template("country.html", country=country, rows=rows, country_id=country_id)
@app.route("/download")
def download():
    import csv
    import io
    from flask import Response
    
    country = request.args.get("country", "VNM")
    indicator_id = request.args.get("indicator", 1)
    file_format = request.args.get("format", "csv")
    
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("""
        SELECT c.country_name, i.indicator_name, t.year_id, f.value, u.unit_name,
               t.is_crisis_year
        FROM fact_economic_data f
        JOIN dim_countries c ON f.country_id = c.country_id
        JOIN dim_indicators i ON f.indicator_id = i.indicator_id
        JOIN dim_time t ON f.year_id = t.year_id
        JOIN dim_units u ON i.unit_id = u.unit_id
        WHERE f.country_id = %s AND f.indicator_id = %s
        ORDER BY t.year_id
    """, (country, indicator_id))
    rows = cur.fetchall()
    cur.close()
    conn.close()

    if file_format == "csv":
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(["Country", "Indicator", "Year", "Value", "Unit", "Crisis Year"])
        for row in rows:
            writer.writerow([row[0], row[1], row[2], row[3], row[4], "Yes" if row[5] else "No"])
        output.seek(0)
        return Response(
            output.getvalue(),
            mimetype="text/csv",
            headers={"Content-Disposition": f"attachment; filename=ASEAN_data.csv"}
        )

    elif file_format == "excel":
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        import io as byte_io
        from flask import send_file

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "ASEAN Economic Data"

        # Header style
        header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        crisis_fill = PatternFill(start_color="FDECEA", end_color="FDECEA", fill_type="solid")

        headers = ["Country", "Indicator", "Year", "Value", "Unit", "Crisis Year"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        # Data rows
        for row_idx, row in enumerate(rows, 2):
            ws.cell(row=row_idx, column=1, value=row[0])
            ws.cell(row=row_idx, column=2, value=row[1])
            ws.cell(row=row_idx, column=3, value=row[2])
            ws.cell(row=row_idx, column=4, value=float(row[3]))
            ws.cell(row=row_idx, column=5, value=row[4])
            crisis_cell = ws.cell(row=row_idx, column=6, value="Yes" if row[5] else "No")
            if row[5]:
                for col in range(1, 7):
                    ws.cell(row=row_idx, column=col).fill = crisis_fill

        # Auto fit columns
        for col in ws.columns:
            max_length = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = max_length + 4

        output = byte_io.BytesIO()
        wb.save(output)
        output.seek(0)
        return send_file(
            output,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name="ASEAN_data.xlsx"
        )
@app.route("/map")
@app.route("/map")
def map_page():
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("""
        SELECT c.country_id, c.country_name, ig.label, r.region_name, fci.year_id
        FROM dim_countries c
        JOIN fact_country_income fci ON c.country_id = fci.country_id
        JOIN dim_income_groups ig ON fci.income_group_id = ig.income_group_id
        JOIN dim_regions r ON c.region_id = r.region_id
        ORDER BY fci.year_id, c.country_name
    """)
    rows = cur.fetchall()
    cur.execute("SELECT DISTINCT year_id FROM fact_country_income ORDER BY year_id")
    years = [r[0] for r in cur.fetchall()]
    cur.close()
    conn.close()

    # Build a dictionary: {year: [{country_id, name, income, region}]}
    data = {}
    for row in rows:
        year = row[4]
        if year not in data:
            data[year] = []
        data[year].append({
            "id": row[0],
            "name": row[1],
            "income": row[2],
            "region": row[3]
        })

    return render_template("map.html", data=data, years=years)
@app.route("/correlation")
def correlation():
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("SELECT country_id, country_name FROM dim_countries ORDER BY country_name")
    countries = cur.fetchall()
    cur.execute("SELECT indicator_id, indicator_name FROM dim_indicators ORDER BY indicator_id")
    indicators = cur.fetchall()
    cur.close()
    conn.close()
    return render_template("correlation.html", countries=countries, indicators=indicators)

@app.route("/correlation/data")
def correlation_data():
    mode = request.args.get("mode", "single")
    country1 = request.args.get("country1", "VNM")
    country2 = request.args.get("country2", "THA")
    indicator1 = request.args.get("indicator1", "GDP")
    indicator2 = request.args.get("indicator2", "Inflation")

    conn = get_connection()
    cur = conn.cursor()

    # Call correlate function directly from PostgreSQL
    cur.execute("""
        SELECT * FROM correlate(%s, %s, %s, %s, %s)
    """, (mode, country1, indicator1, indicator2,
          country2 if mode == 'compare' else None))
    row = cur.fetchone()

    # Get scatter plot points
    if mode == "single":
        cur.execute("""
            SELECT f1.year_id, f1.value, f2.value
            FROM fact_economic_data f1
            JOIN fact_economic_data f2
                ON f1.year_id = f2.year_id AND f1.country_id = f2.country_id
            JOIN dim_indicators i1 ON f1.indicator_id = i1.indicator_id
            JOIN dim_indicators i2 ON f2.indicator_id = i2.indicator_id
            WHERE f1.country_id = %s
            AND i1.indicator_name ILIKE %s
            AND i2.indicator_name ILIKE %s
            ORDER BY f1.year_id
        """, (country1, f'%{indicator1}%', f'%{indicator2}%'))
    else:
        cur.execute("""
            SELECT f1.year_id, f1.value, f2.value
            FROM fact_economic_data f1
            JOIN fact_economic_data f2 ON f1.year_id = f2.year_id
            JOIN dim_indicators i1 ON f1.indicator_id = i1.indicator_id
            JOIN dim_indicators i2 ON f2.indicator_id = i2.indicator_id
            WHERE f1.country_id = %s AND f2.country_id = %s
            AND i1.indicator_name ILIKE %s
            AND i2.indicator_name ILIKE %s
            ORDER BY f1.year_id
        """, (country1, country2, f'%{indicator1}%', f'%{indicator2}%'))

    points = cur.fetchall()
    cur.close()
    conn.close()

    if not row or row[0] is None:
        return jsonify({"error": "Insufficient data"})

    return jsonify({
        "corr": float(row[0]),
        "interpretation": row[1],
        "obs": row[2],
        "name1": row[3],
        "name2": row[4],
        "points": [{"year": p[0], "x": float(p[1]), "y": float(p[2])} for p in points]
    })
if __name__ == "__main__":
    app.run(debug=True)